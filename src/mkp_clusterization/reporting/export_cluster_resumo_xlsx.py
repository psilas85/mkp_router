# =========================================================
# 📦 src/mkp_clusterization/reporting/export_cluster_resumo_xlsx.py
# =========================================================

import os
import argparse
import pandas as pd
from loguru import logger
from database.db_connection import get_connection
from mkp_clusterization.domain.haversine_utils import haversine
from mkp_clusterization.domain.reverse_geocode_utils import enrich_centros_reverse

VELOCIDADE_MEDIA_KMH = 30.0  # regra executiva fixa


# =========================================================
# 🧮 Métricas geométricas por cluster
# =========================================================
def calcular_metricas_por_cluster(df_pdvs: pd.DataFrame) -> pd.DataFrame:
    registros = []

    for cluster_label, g in df_pdvs.groupby("cluster_label"):
        distancias = []

        for _, row in g.iterrows():
            d = haversine(
                (row["pdv_lat"], row["pdv_lon"]),
                (row["centro_lat"], row["centro_lon"]),
            )
            distancias.append(d)

        if not distancias:
            continue

        dist_media = sum(distancias) / len(distancias)
        dist_max = max(distancias)

        tempo_medio = (dist_media / VELOCIDADE_MEDIA_KMH) * 60
        tempo_max = (dist_max / VELOCIDADE_MEDIA_KMH) * 60

        registros.append(
            {
                "cluster_label": cluster_label,
                "enderecos": g["pdv_id"].nunique(),
                "qde_clientes": g["pdv_vendas"].sum(),
                "dist_media_km": round(dist_media, 2),
                "dist_max_km": round(dist_max, 2),
                "tempo_medio_min": round(tempo_medio, 1),
                "tempo_max_min": round(tempo_max, 1),
            }
        )

    return pd.DataFrame(registros)



# =========================================================
# 📤 Exportação principal
# =========================================================
def exportar_cluster_resumo(tenant_id: int, clusterization_id: str):
    logger.info(
        f"📊 Exportando resumo de clusters | tenant={tenant_id} | clusterization_id={clusterization_id}"
    )

    conn = get_connection()

    # =========================================================
    # 🔍 Run mais recente
    # =========================================================
    query_run = f"""
        SELECT id AS run_id
        FROM cluster_run
        WHERE tenant_id = {tenant_id}
          AND clusterization_id = '{clusterization_id}'
        ORDER BY criado_em DESC
        LIMIT 1;
    """
    run_df = pd.read_sql_query(query_run, conn)

    if run_df.empty:
        conn.close()
        raise ValueError("❌ Nenhum run encontrado")

    run_id = int(run_df.iloc[0]["run_id"])

    # =========================================================
    # 📊 View resumo (base executiva)
    # =========================================================
    query_resumo = f"""
        SELECT *
        FROM v_cluster_resumo
        WHERE tenant_id = {tenant_id}
          AND run_id = {run_id}
        ORDER BY cluster_label;
    """
    df = pd.read_sql_query(query_resumo, conn)

    if df.empty:
        conn.close()
        raise ValueError("❌ Nenhum dado encontrado em v_cluster_resumo")

    # =========================================================
    # 📍 PDVs por cluster (para cálculo geométrico)
    # =========================================================
    query_pdvs = f"""
        SELECT
            cs.cluster_label,
            cs.centro_lat,
            cs.centro_lon,
            p.id AS pdv_id,
            p.pdv_lat,
            p.pdv_lon,
            COALESCE(p.pdv_vendas, 0) AS pdv_vendas
        FROM cluster_setor cs
        JOIN cluster_setor_pdv csp
            ON csp.cluster_id = cs.id
           AND csp.tenant_id = cs.tenant_id
        JOIN pdvs p
            ON p.id = csp.pdv_id
           AND p.tenant_id = cs.tenant_id
        WHERE cs.run_id = {run_id}
          AND cs.tenant_id = {tenant_id}
          AND p.pdv_lat IS NOT NULL
          AND p.pdv_lon IS NOT NULL;
    """
    df_pdvs = pd.read_sql_query(query_pdvs, conn)
    conn.close()

    if df_pdvs.empty:
        raise ValueError("❌ Nenhum PDV encontrado para cálculo de métricas")

    # =========================================================
    # 🧮 Cálculo geométrico
    # =========================================================
    df_calc = calcular_metricas_por_cluster(df_pdvs)

    df = df.merge(df_calc, on="cluster_label", how="left", suffixes=("", "_calc"))

    for col in [
        "dist_media_km",
        "dist_max_km",
        "tempo_medio_min",
        "tempo_max_min",
    ]:
        df[col] = df[f"{col}_calc"].fillna(df[col])

    df = df.drop(columns=[c for c in df.columns if c.endswith("_calc")], errors="ignore")

    # =========================================================
    # 🧹 Limpeza técnica
    # =========================================================
    df = df.drop(columns=["metrics_json"], errors="ignore")

    # =========================================================
    # 🧭 Reverse geocode dos centros
    # =========================================================
    df_geo = enrich_centros_reverse(df)
    df = df.merge(df_geo, on=["centro_lat", "centro_lon"], how="left")
    df = df.rename(
        columns={
            "Endereco centro": "Logradouro / Bairro",
            "Cidade centro": "Cidade",
            "UF centro": "UF",
            "CEP centro": "CEP",
        }
    )


    # =========================================================
    # 🏷️ Renomeio executivo
    # =========================================================
    df = df.rename(
        columns={
            "cluster_label": "Cluster",
            "enderecos": "Endereços",
            "qde_clientes": "Qde de Clientes",
            "dist_media_km": "Distância média (km)",
            "dist_max_km": "Distância máxima (km)",
            "tempo_medio_min": "Tempo médio (min)",
            "tempo_max_min": "Tempo máximo (min)",
            "centro_lat": "Latitude centro",
            "centro_lon": "Longitude centro",
        }
    )


    # =========================================================
    # 📐 Ordem final (CEP removido)
    # =========================================================
    df = df.drop(columns=["CEP"], errors="ignore")

    df = df[
        [
            "Cluster",
            "Endereços",
            "Qde de Clientes",
            "Distância média (km)",
            "Distância máxima (km)",
            "Tempo médio (min)",
            "Tempo máximo (min)",
            "Logradouro / Bairro",
            "Cidade",
            "UF",
            "Latitude centro",
            "Longitude centro",
        ]
    ]

    # =========================================================
    # 📤 Exportação Excel
    # =========================================================
    output_dir = f"output/reports/{tenant_id}"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(
        output_dir, f"cluster_resumo_{clusterization_id}.xlsx"
    )

    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Resumo por Cluster", index=False)
        ws = writer.book["Resumo por Cluster"]

        ws.freeze_panes = "A2"

        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align

        widths = {
            1: 10,
            2: 8,
            3: 10,
            4: 22,
            5: 24,
            6: 20,
            7: 20,
            8: 40,  # Logradouro / Bairro
            9: 20,  # Cidade
            10: 6,  # UF
            11: 18,
            12: 18,
        }

        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    logger.success(f"✅ Excel executivo gerado: {output_path}")


# =========================================================
# 🚀 CLI
# =========================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--tenant_id", type=int, required=True)
    parser.add_argument("--clusterization_id", type=str, required=True)
    args = parser.parse_args()

    exportar_cluster_resumo(args.tenant_id, args.clusterization_id)
